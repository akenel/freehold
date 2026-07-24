"""Freehold — outbound formats. The right-hand side of the service interface.

A report is stored as JSON because JSON is the honest wire format. But nobody
asked for their customer list back as JSON — they asked for it back *cleaned*, in
the shape they sent it. xlsx in, xlsx out. So this turns one stored report into
whatever a human actually wants to open:

    xlsx   a real workbook: frozen bold header, autofilter, a coloured Gate
           column, the enriched values folded in, and two extra tabs — the review
           queue on its own, and the run's provenance. This is the deliverable.
    csv    the same grid, flat, for the system on the other end that wants a feed.
    json   the raw report, untouched — for diffing and re-import.

The rule the user asked for: **default to the format they gave us.** A spreadsheet
came in, a spreadsheet goes out. The reader (inbound.py) is stdlib and paranoid
because it parses untrusted files; the writer here is openpyxl because we control
every byte it writes — there is no bomb to build out of our own records.
"""
import csv
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# One value per cell for the flat exports: the enriched fields carry a whole
# provenance envelope, but a spreadsheet cell wants the answer, not the paperwork.
_GATE_FILL = {
    "auto": "C7EBD1",       # green — applied
    "review": "F6E2B8",     # amber — a human should look
    "rejected": "F5C6C6",   # red — held back
}
_HEAD_FILL = "2C3E50"
_HEAD_FONT = "FFFFFF"


def _cell_value(v):
    """Flatten one enriched value to something a cell can hold.

    quality -> its score; reconcile -> its status; everything else -> as-is. The
    detail (what was missing, which numbers disagreed) is a column of its own."""
    if isinstance(v, dict):
        if "score" in v:
            return v["score"]
        if "status" in v:
            return v["status"]
        return json.dumps(v, ensure_ascii=False)
    return v


def _detail(v):
    """The human note that goes beside a flattened value, or ""."""
    if isinstance(v, dict):
        if "missing" in v and v.get("missing"):
            return "missing: " + ", ".join(v["missing"])
        if "status" in v:
            bits = []
            for k, nums in v.items():
                if isinstance(nums, list) and nums:
                    bits.append(f"{k}: {', '.join(nums)}")
            return " | ".join(bits)
    return ""


def clean_grid(report: dict) -> tuple[list[dict], list[dict]]:
    """Fold each cleaned value back into the column it cleans. The shared shape
    every format and the preview render from.

    Returns (columns, rows).
      columns: [{"name", "kind"}] where kind is
               "source"  — an original column, showing the CLEAN value when one
                           exists (the whole point: the phone column becomes
                           dialable instead of the fix living in a sidecar)
               "raw"     — the original of a cleaned column, kept beside it, so
                           nothing is destroyed (principle 3)
               "enriched"— an added analysis column (quality, phone_check) that
                           replaces nothing
      rows: [{name: {"value", "raw", "changed", "source", "confidence"}}], plus
            "_gate" and "_label". `changed` marks a cell the pipeline improved."""
    meta = report.get("meta", {})
    records = report.get("records", [])

    source = list(meta.get("columns") or [])
    replaced: dict[str, str] = {}      # source column -> the enriched field that cleans it
    added: list[str] = []              # enriched fields that replace nothing
    for r in records:
        for k in r:
            if not k.startswith("_") and k not in source:
                source.append(k)
        for k, env in (r.get("_enriched") or {}).items():
            rep = env.get("replaces")
            if rep:
                replaced.setdefault(rep, k)
            elif k not in added:
                added.append(k)

    columns: list[dict] = []
    for c in source:
        columns.append({"name": c, "kind": "source"})
        if c in replaced:
            columns.append({"name": f"{c} · original", "kind": "raw"})
    for e in added:
        columns.append({"name": e, "kind": "enriched"})

    rows = []
    for r in records:
        enr = r.get("_enriched") or {}
        cells: dict[str, dict] = {}
        for c in source:
            if c in replaced and replaced[c] in enr:
                env = enr[replaced[c]]
                clean = _cell_value(env.get("value"))
                raw = r.get(c, "")
                changed = str(clean) not in ("", "None") and str(clean) != str(raw)
                cells[c] = {"value": clean if str(clean) not in ("", "None") else raw,
                            "raw": raw, "changed": changed,
                            "source": env.get("source"), "confidence": env.get("confidence")}
                cells[f"{c} · original"] = {"value": raw, "changed": False}
            else:
                cells[c] = {"value": r.get(c, ""), "changed": False}
                if c in replaced:      # cleaned on other rows, declined on this one
                    cells[f"{c} · original"] = {"value": r.get(c, ""), "changed": False}
        for e in added:
            env = enr.get(e)
            if env is None:
                cells[e] = {"value": "", "changed": False}
            else:
                val = _cell_value(env.get("value"))
                note = _detail(env.get("value"))
                cells[e] = {"value": f"{val}  ({note})" if note else val, "changed": False,
                            "source": env.get("source"), "confidence": env.get("confidence")}
        cells["_gate"] = r.get("_gate", "")
        cells["_label"] = r.get("_label", "")
        rows.append(cells)
    return columns, rows


# --- csv -------------------------------------------------------------------
def to_csv(report: dict) -> bytes:
    columns, rows = clean_grid(report)
    header = [c["name"] for c in columns] + ["Gate"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    for row in rows:
        w.writerow([row.get(c["name"], {}).get("value", "") for c in columns]
                   + [row.get("_gate", "")])
    return buf.getvalue().encode("utf-8-sig")   # BOM so Excel opens UTF-8 cleanly


# --- json ------------------------------------------------------------------
def to_json(report: dict) -> bytes:
    return json.dumps(report, indent=2, ensure_ascii=False).encode("utf-8")


# --- xlsx ------------------------------------------------------------------
_CHG_FILL = "EDE4FB"          # a cleaned cell — faint purple, so it reads at a glance
_RAW_FONT = "8A8A8A"          # the kept-original columns, greyed


def _write_sheet(ws, columns, rows):
    head_fill = PatternFill("solid", fgColor=_HEAD_FILL)
    head_font = Font(bold=True, color=_HEAD_FONT)
    gate_fills = {g: PatternFill("solid", fgColor=c) for g, c in _GATE_FILL.items()}
    chg_fill = PatternFill("solid", fgColor=_CHG_FILL)
    raw_font = Font(color=_RAW_FONT, italic=True)
    header = [c["name"] for c in columns] + ["Gate"]
    ncol = len(header)

    for j, name in enumerate(header, start=1):
        c = ws.cell(row=1, column=j, value=name)
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(vertical="center")

    widths = [max(10, min(48, len(str(h)) + 2)) for h in header]
    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(columns, start=1):
            cell = row.get(col["name"], {})
            val = cell.get("value", "")
            c = ws.cell(row=i, column=j, value=val)
            if col["kind"] == "raw":
                c.font = raw_font
            elif cell.get("changed"):
                c.fill = chg_fill                          # this cell got cleaner
            widths[j - 1] = max(widths[j - 1], min(48, len(str(val)) + 2))
        g = row.get("_gate", "")
        gc = ws.cell(row=i, column=ncol, value=g)
        if gate_fills.get(g):
            gc.fill = gate_fills[g]

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"                                  # header stays put
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{len(rows) + 1}"


def to_xlsx(report: dict) -> bytes:
    """The deliverable: a workbook better than the one that came in."""
    columns, rows = clean_grid(report)
    meta = report.get("meta", {})
    wb = Workbook()

    ws = wb.active
    ws.title = "Cleaned"
    _write_sheet(ws, columns, rows)

    # The review queue on its own tab — the rows a human actually has to touch,
    # which is the whole point of the gate. Empty tab if the run was all-green.
    review = [r for r in rows if r.get("_gate") == "review"]
    rq = wb.create_sheet("Needs review")
    if review:
        _write_sheet(rq, columns, review)
    else:
        rq["A1"] = "Nothing needed review — every row cleared the gate."

    # Provenance, so the file stands alone once it leaves the box.
    info = wb.create_sheet("How this was made")
    facts = [
        ("Source", meta.get("source", "")),
        ("Run at", meta.get("run_at", "")),
        ("Run by", meta.get("run_by", "")),
        ("Records", meta.get("count", "")),
        ("Brain", meta.get("model", "none")),
        ("Prompt version", meta.get("prompt_version", "")),
        ("AI-written rows", meta.get("enriched_count", 0)),
        ("Tokens", meta.get("tokens", 0)),
        ("Needs review", meta.get("review_count", 0)),
        ("", ""),
        ("Findings", ""),
    ]
    for f in meta.get("findings", []):
        facts.append(("", f.get("text", "")))
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    for i, (k, v) in enumerate(facts, start=1):
        info.cell(row=i, column=1, value=k).font = Font(bold=True)
        info.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- dispatch --------------------------------------------------------------
FORMATS = {
    "xlsx": (to_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "csv": (to_csv, "text/csv; charset=utf-8"),
    "json": (to_json, "application/json"),
}


def default_format(report: dict) -> str:
    """Give them back what they gave us. A spreadsheet came in → a spreadsheet
    goes out. Only the API-channel demo (no uploaded file) defaults to json."""
    return "xlsx" if report.get("meta", {}).get("channel") in ("spreadsheet", "list") else "json"


def render(report: dict, fmt: str) -> tuple[bytes, str]:
    fn, mime = FORMATS[fmt]
    return fn(report), mime
